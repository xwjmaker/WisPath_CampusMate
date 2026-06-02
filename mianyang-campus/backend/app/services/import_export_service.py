import io
import re
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.models.user import User, UserRole
from app.core.security import hash_password


DEFAULT_PASSWORD = "123456"

# 学号/工号格式：纯数字或字母数字组合，4-20位
USERNAME_PATTERN = re.compile(r'^[A-Za-z0-9]{4,20}$')
# 手机号格式：11位数字
PHONE_PATTERN = re.compile(r'^1[3-9]\d{9}$')


def validate_username(username: str) -> str | None:
    """校验用户名格式，返回错误信息或 None"""
    if not username:
        return "学号/工号不能为空"
    if not USERNAME_PATTERN.match(username):
        return f"学号/工号格式无效: '{username}'（应为4-20位字母或数字）"
    return None


def validate_name(name: str) -> str | None:
    """校验姓名格式"""
    if not name:
        return "姓名不能为空"
    if len(name) > 50:
        return f"姓名过长: '{name}'（最多50个字符）"
    return None


def validate_phone(phone: str | None) -> str | None:
    """校验手机号格式"""
    if phone and not PHONE_PATTERN.match(phone):
        return f"手机号格式无效: '{phone}'（应为11位有效手机号）"
    return None


def validate_age(age: int | None, row_num: int) -> tuple[int | None, str | None]:
    """校验年龄，返回 (处理后的年龄, 错误信息)"""
    if age is None:
        return None, None
    if not isinstance(age, (int, float)):
        try:
            age = int(age)
        except (ValueError, TypeError):
            return None, f"年龄格式无效: '{age}'"
    age = int(age)
    if age < 1 or age > 150:
        return None, f"年龄超出合理范围: {age}（应为1-150）"
    return age, None


def export_users(db: Session, role: UserRole) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用户数据"

    if role == UserRole.STUDENT:
        headers = ["学号", "姓名", "学院", "班级", "性别", "年龄", "联系电话", "籍贯"]
        ws.append(headers)
        users = db.query(User).filter(User.role == UserRole.STUDENT).all()
        for u in users:
            ws.append([
                u.username,
                u.name,
                u.college or "",
                u.class_name or "",
                u.gender or "",
                u.age or "",
                u.phone or "",
                u.hometown or "",
            ])
    else:
        headers = ["工号", "姓名", "学院", "性别", "年龄", "职称", "所属单位", "联系电话"]
        ws.append(headers)
        users = db.query(User).filter(User.role == UserRole.TEACHER).all()
        for u in users:
            ws.append([
                u.username,
                u.name,
                u.college or "",
                u.gender or "",
                u.age or "",
                u.title or "",
                u.department or "",
                u.phone or "",
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def import_users(db: Session, role: UserRole, file_content: bytes) -> dict:
    """导入用户数据，支持去重、格式校验和数据清洗"""
    wb = load_workbook(io.BytesIO(file_content))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return {"total": 0, "created": 0, "skipped": 0, "errors": []}

    created = 0
    skipped = 0
    errors = []

    # 预加载已有用户名，用于批量去重（避免逐行查库）
    existing_usernames = {
        u.username for u in db.query(User.username).filter(User.role == role).all()
    }
    # 文件内去重：同一文件中可能出现重复用户名
    file_usernames_seen: set[str] = set()

    for i, row in enumerate(rows, start=2):
        try:
            if role == UserRole.STUDENT:
                if len(row) < 2:
                    errors.append(f"第{i}行: 数据不完整（至少需要学号和姓名）")
                    continue
                username = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if row[1] else ""
                college = str(row[2]).strip() if len(row) > 2 and row[2] else None
                class_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
                gender = str(row[4]).strip() if len(row) > 4 and row[4] else None
                age = row[5] if len(row) > 5 and row[5] else None
                phone = str(row[6]).strip() if len(row) > 6 and row[6] else None
                hometown = str(row[7]).strip() if len(row) > 7 and row[7] else None
            else:
                if len(row) < 2:
                    errors.append(f"第{i}行: 数据不完整（至少需要工号和姓名）")
                    continue
                username = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if row[1] else ""
                college = str(row[2]).strip() if len(row) > 2 and row[2] else None
                gender = str(row[3]).strip() if len(row) > 3 and row[3] else None
                age = row[4] if len(row) > 4 and row[4] else None
                title = str(row[5]).strip() if len(row) > 5 and row[5] else None
                department = str(row[6]).strip() if len(row) > 6 and row[6] else None
                phone = str(row[7]).strip() if len(row) > 7 and row[7] else None

            # ====== 格式校验 ======
            err = validate_username(username)
            if err:
                errors.append(f"第{i}行: {err}")
                continue

            err = validate_name(name)
            if err:
                errors.append(f"第{i}行: {err}")
                continue

            err = validate_phone(phone)
            if err:
                errors.append(f"第{i}行: {err}")
                continue

            age, err = validate_age(age, i)
            if err:
                errors.append(f"第{i}行: {err}")
                continue

            # ====== 去重检查 ======
            if username in existing_usernames:
                skipped += 1
                continue
            if username in file_usernames_seen:
                errors.append(f"第{i}行: 文件内重复学号/工号 '{username}'，已跳过")
                continue

            # ====== 数据清洗 ======
            # 清除空白符
            name = re.sub(r'\s+', '', name)
            if college:
                college = college.strip()
            if class_name:
                class_name = class_name.strip()
            if gender and gender not in ("男", "女"):
                errors.append(f"第{i}行: 性别值无效: '{gender}'（应为'男'或'女'）")
                continue
            if phone:
                phone = phone.strip()

            file_usernames_seen.add(username)

            user = User(
                username=username,
                password_hash=hash_password(DEFAULT_PASSWORD),
                name=name,
                role=role,
                college=college,
                gender=gender,
                age=age,
                phone=phone,
            )

            if role == UserRole.STUDENT:
                user.hometown = hometown
                user.class_name = class_name
            else:
                user.title = title
                user.department = department

            db.add(user)
            existing_usernames.add(username)
            created += 1

        except Exception as e:
            errors.append(f"第{i}行: 处理异常 - {str(e)}")

    db.commit()

    return {
        "total": len(rows),
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
